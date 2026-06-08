import io
from datetime import date, timedelta

from django.db.models import Count, Sum
from django.db.models.functions import TruncWeek
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.authentication.permissions import allow_roles
from apps.inventario.models import Lote, MovimientoInventario
from apps.produccion.models import Batido, LoteProductoTerminado
from apps.recepcion.models import OrdenCompra, RecepcionMercancia

_IND_ROLES = ('ADMIN', 'GERENTE')


class KpisView(APIView):
    permission_classes = [allow_roles(*_IND_ROLES)]

    def get(self, request):
        tipo = request.query_params.get('tipo', 'stock')

        if tipo == 'stock':
            return self._stock()
        if tipo == 'vencimientos':
            dias = int(request.query_params.get('dias', 7))
            return self._vencimientos(dias)
        if tipo == 'produccion':
            desde = request.query_params.get('desde')
            hasta = request.query_params.get('hasta')
            return self._produccion(desde, hasta)
        return Response({'detail': 'tipo inválido. Use: stock, vencimientos, produccion'}, status=400)

    def _stock(self):
        rows = (
            Lote.objects.values(
                'bodega__nombre',
                'bodega__tipo',
                'materia_prima__nombre',
                'materia_prima__unidad_medida__simbolo',
            )
            .annotate(cantidad_total=Sum('cantidad'))
            .order_by('bodega__nombre', '-cantidad_total')
        )
        stock_por_bodega = [
            {
                'bodega_nombre':  r['bodega__nombre'],
                'bodega_tipo':    r['bodega__tipo'],
                'materia_prima':  r['materia_prima__nombre'],
                'unidad':         r['materia_prima__unidad_medida__simbolo'],
                'cantidad_total': str(r['cantidad_total']),
            }
            for r in rows
        ]
        return Response({'stock_por_bodega': stock_por_bodega})

    def _vencimientos(self, dias):
        hoy = date.today()
        limite = hoy + timedelta(days=dias)
        lotes = (
            Lote.objects
            .filter(fecha_vencimiento__range=(hoy, limite))
            .select_related('materia_prima', 'bodega')
            .order_by('fecha_vencimiento')
        )
        lotes_por_vencer = [
            {
                'lote_id':           lote.id,
                'numero_lote':       lote.numero_lote,
                'materia_prima':     lote.materia_prima.nombre,
                'bodega':            lote.bodega.nombre,
                'cantidad':          str(lote.cantidad),
                'fecha_vencimiento': str(lote.fecha_vencimiento),
                'dias_restantes':    (lote.fecha_vencimiento - hoy).days,
            }
            for lote in lotes
        ]
        return Response({'lotes_por_vencer': lotes_por_vencer})

    def _produccion(self, desde, hasta):
        qs = LoteProductoTerminado.objects.all()
        if desde:
            qs = qs.filter(fecha_produccion__gte=desde)
        if hasta:
            qs = qs.filter(fecha_produccion__lte=hasta)
        total = qs.aggregate(total=Sum('cantidad'))['total'] or 0
        return Response({'total_unidades_producidas': total})


class ExportarView(APIView):
    permission_classes = [allow_roles(*_IND_ROLES)]

    def get(self, request):
        formato = request.query_params.get('formato', 'xlsx')
        if formato == 'pdf':
            return self._exportar_pdf()
        if formato == 'xlsx':
            return self._exportar_excel()
        return Response({'detail': 'formato inválido. Use: pdf, xlsx'}, status=400)

    def _stock_data(self):
        return (
            Lote.objects.values(
                'bodega__nombre',
                'bodega__tipo',
                'materia_prima__nombre',
                'materia_prima__unidad_medida__simbolo',
            )
            .annotate(cantidad_total=Sum('cantidad'))
            .order_by('bodega__nombre', '-cantidad_total')
        )

    def _exportar_excel(self):
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Inventario'

        headers = ['Bodega', 'Tipo', 'Materia Prima', 'Unidad', 'Cantidad Total']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in self._stock_data():
            ws.append([
                row['bodega__nombre'],
                row['bodega__tipo'],
                row['materia_prima__nombre'],
                row['materia_prima__unidad_medida__simbolo'],
                float(row['cantidad_total']),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'inventario_{date.today()}.xlsx'
        response = HttpResponse(buf.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _exportar_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('Reporte de Inventario — Daluzed', styles['Heading1']))
        elements.append(Paragraph(f'Generado: {date.today()}', styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        rows = [['Bodega', 'Tipo', 'Materia Prima', 'Unidad', 'Cantidad']]
        for row in self._stock_data():
            rows.append([
                row['bodega__nombre'],
                row['bodega__tipo'],
                row['materia_prima__nombre'],
                row['materia_prima__unidad_medida__simbolo'] or '',
                str(row['cantidad_total']),
            ])

        t = Table(rows, hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B1A1A')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5EE')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDBBAA')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)

        filename = f'inventario_{date.today()}.pdf'
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class UtilizacionBodegaView(APIView):
    permission_classes = [allow_roles(*_IND_ROLES)]

    def get(self, request):
        from decimal import Decimal
        from apps.inventario.models import ZonaBodega
        from django.db.models import Sum, Value
        from django.db.models.functions import Coalesce

        zonas = (
            ZonaBodega.objects
            .select_related('bodega')
            .annotate(
                stock_actual=Coalesce(Sum('lotes__cantidad'), Value(Decimal('0')))
            )
            .values(
                'id',
                'nombre',
                'bodega__nombre',
                'bodega__tipo',
                'capacidad_maxima',
                'stock_actual',
            )
        )

        data = []
        for zona in zonas:
            capacidad = float(zona['capacidad_maxima']) or 1
            stock = float(zona['stock_actual']) or 0
            porcentaje = min(100, (stock / capacidad * 100)) if capacidad > 0 else 0

            data.append({
                'zona_id': zona['id'],
                'zona_nombre': zona['nombre'],
                'bodega_nombre': zona['bodega__nombre'],
                'bodega_tipo': zona['bodega__tipo'],
                'stock_actual': str(stock),
                'capacidad_maxima': str(zona['capacidad_maxima']),
                'porcentaje_utilizacion': round(porcentaje, 1),
            })

        return Response({
            'utilizacion_por_zona': sorted(data, key=lambda x: x['bodega_nombre'])
        })


class ReporteSemanalView(APIView):
    permission_classes = [allow_roles(*_IND_ROLES)]

    def get(self, request):
        try:
            hasta: date = date.fromisoformat(request.query_params['hasta']) if 'hasta' in request.query_params else date.today()
            desde: date = date.fromisoformat(request.query_params['desde']) if 'desde' in request.query_params else hasta - timedelta(weeks=4)
        except ValueError:
            return Response({'detail': 'Formato de fecha inválido. Use YYYY-MM-DD.'}, status=400)

        semanas = self._agrupar_por_semana(desde, hasta)

        if request.query_params.get('formato') == 'xlsx':
            return self._exportar_excel(semanas, desde, hasta)

        return Response({'desde': str(desde), 'hasta': str(hasta), 'semanas': semanas})

    def _agrupar_por_semana(self, desde: date, hasta: date) -> list[dict]:
        batidos_qs = (
            Batido.objects
            .filter(fecha_produccion__range=(desde, hasta))
            .annotate(semana=TruncWeek('fecha_produccion'))
            .values('semana')
            .annotate(total=Count('id'))
            .order_by('semana')
        )
        recepciones_qs = (
            RecepcionMercancia.objects
            .filter(fecha__range=(desde, hasta))
            .annotate(semana=TruncWeek('fecha'))
            .values('semana')
            .annotate(total=Count('id'))
            .order_by('semana')
        )
        despachos_qs = (
            LoteProductoTerminado.objects
            .filter(fecha_despacho__isnull=False, fecha_despacho__range=(desde, hasta))
            .annotate(semana=TruncWeek('fecha_despacho'))
            .values('semana')
            .annotate(total=Count('id'), unidades=Sum('cantidad'))
            .order_by('semana')
        )

        data: dict[str, dict] = {}

        def _key(semana_dt) -> str:
            return semana_dt.strftime('%Y-%m-%d')

        def _ensure(k: str) -> None:
            data.setdefault(k, {
                'semana_inicio': k,
                'batidos': 0,
                'recepciones': 0,
                'despachos': 0,
                'unidades_despachadas': '0',
            })

        for row in batidos_qs:
            k = _key(row['semana'])
            _ensure(k)
            data[k]['batidos'] = row['total']

        for row in recepciones_qs:
            k = _key(row['semana'])
            _ensure(k)
            data[k]['recepciones'] = row['total']

        for row in despachos_qs:
            k = _key(row['semana'])
            _ensure(k)
            data[k]['despachos'] = row['total']
            data[k]['unidades_despachadas'] = str(row['unidades'] or 0)

        return sorted(data.values(), key=lambda x: x['semana_inicio'])

    def _exportar_excel(self, semanas: list[dict], desde: date, hasta: date) -> HttpResponse:  # noqa: E501
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte Semanal'

        ws.append([f'Reporte semanal Daluzed — {desde} al {hasta}'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])

        headers = ['Semana (inicio lunes)', 'Batidos', 'Recepciones', 'Despachos', 'Unidades despachadas']
        ws.append(headers)
        for cell in ws[3]:
            cell.font = Font(bold=True)

        for sem in semanas:
            ws.append([
                sem['semana_inicio'],
                sem['batidos'],
                sem['recepciones'],
                sem['despachos'],
                float(sem['unidades_despachadas']),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'reporte_semanal_{desde}_{hasta}.xlsx'
        response = HttpResponse(buf.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ResumenView(APIView):
    """
    GET /api/v1/indicadores/resumen/
    Resumen consolidado para el Dashboard gerencial.
    Retorna stock BP total, rotación del último mes, batidos de la semana,
    lotes próximos a vencer y OC pendientes/parciales.
    RF-IND-01, RF-IND-07
    """
    permission_classes = [allow_roles(*_IND_ROLES)]

    def get(self, request):
        hoy = date.today()
        hace_30_dias = hoy - timedelta(days=30)

        # ── Stock Bodega Principal ────────────────────────────────────
        stock_bp_total = (
            Lote.objects
            .filter(bodega__tipo='PRINCIPAL')
            .aggregate(total=Sum('cantidad'))['total'] or 0
        )

        # ── Rotación de inventario (último mes) ───────────────────────
        consumo_mes = (
            MovimientoInventario.objects
            .filter(tipo='CONSUMO', fecha__date__gte=hace_30_dias)
            .aggregate(total=Sum('cantidad'))['total'] or 0
        )
        stock_total_actual = (
            Lote.objects.aggregate(total=Sum('cantidad'))['total'] or 0
        )
        rotacion = (float(consumo_mes) / float(stock_total_actual)) if stock_total_actual > 0 else 0

        # ── Batidos de la semana actual (lunes → hoy) ────────────────
        lunes = hoy - timedelta(days=hoy.weekday())
        DIAS = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sá', 'Do']
        batidos_qs = (
            Batido.objects
            .filter(fecha_produccion__range=(lunes, hoy))
            .values('fecha_produccion')
            .annotate(total=Count('id'))
        )
        batidos_por_fecha: dict[str, int] = {
            str(row['fecha_produccion']): row['total'] for row in batidos_qs
        }
        batidos_semana = []
        for i in range(7):
            d = lunes + timedelta(days=i)
            batidos_semana.append({
                'dia': DIAS[i],
                'fecha': str(d),
                'batidos': batidos_por_fecha.get(str(d), 0),
            })

        # ── Lotes próximos a vencer (7 días) ──────────────────────────
        limite_venc = hoy + timedelta(days=7)
        lotes_por_vencer_count = Lote.objects.filter(
            fecha_vencimiento__range=(hoy, limite_venc)
        ).count()

        # ── OC pendientes o parciales ────────────────────────────────
        oc_pendientes = OrdenCompra.objects.filter(
            estado__in=['PENDIENTE', 'PARCIAL']
        ).count()

        return Response({
            'stock_bp_total': str(stock_bp_total),
            'rotacion_inventario': round(rotacion, 4),
            'periodo_rotacion': f'{hace_30_dias} al {hoy}',
            'batidos_semana': batidos_semana,
            'lotes_por_vencer_count': lotes_por_vencer_count,
            'oc_pendientes': oc_pendientes,
        })
